import time

import os
from datetime import datetime
from PyQt5 import QtCore, QtGui, QtWidgets
import sqlite3
import openpyxl
baglanti = sqlite3.connect("./data.db")
cunsor = baglanti.cursor()
cunsor.execute("CREATE TABLE IF NOT EXISTS ana (isadi TEXT , firma TEXT , avansialan TEXT , verilen int )")
cunsor.execute("CREATE TABLE IF NOT EXISTS FİRMA (id İNT ,tarih TEXT , belge TEXT , FİRMA TEXT , Kişi int , gider TEXT ,toplam int )")
baglanti.commit()
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(629, 303)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(":/icon/excel.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        MainWindow.setStyleSheet("#centralwidget{\n"
"background-color: rgb(172, 255, 200);}")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.widget = QtWidgets.QWidget(self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(20, 40, 581, 201))
        self.widget.setObjectName("widget")
        self.pushButton = QtWidgets.QPushButton(self.widget)
        self.pushButton.setGeometry(QtCore.QRect(300, 119, 87, 43))
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap(":/icon/note.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton.setIcon(icon1)
        self.pushButton.setIconSize(QtCore.QSize(22, 23))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.ileri)
        self.pushButton_2 = QtWidgets.QPushButton(self.widget)
        self.pushButton_2.setGeometry(QtCore.QRect(470, 150, 111, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap(":/icon/exit.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_2.setIcon(icon2)
        self.pushButton_2.setIconSize(QtCore.QSize(25, 25))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.clicked.connect(self.exit)
        self.lineEdit = QtWidgets.QLineEdit(self.widget)
        self.lineEdit.setGeometry(QtCore.QRect(10, 120, 291, 41))
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit.setFont(font)
        self.lineEdit.setText("")
        self.lineEdit.setObjectName("lineEdit")
        self.label = QtWidgets.QLabel(self.widget)
        self.label.setGeometry(QtCore.QRect(20, 60, 292, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.pushButton_3 = QtWidgets.QPushButton(self.widget)
        self.pushButton_3.setGeometry(QtCore.QRect(470, 90, 111, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setIcon(icon)
        self.pushButton_3.setIconSize(QtCore.QSize(24, 23))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_3.clicked.connect(self.kaydet)
        self.label_2 = QtWidgets.QLabel(self.widget)
        self.label_2.setGeometry(QtCore.QRect(0, 0, 581, 21))
        self.label_2.setStyleSheet("#label_2{\n"
"\n"
"\n"
"background-image: url(:/icon/MERT.PNG);}")
        self.label_2.setText("")
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.widget)
        self.label_3.setGeometry(QtCore.QRect(0, 30, 581, 21))
        font = QtGui.QFont()
        font.setPointSize(13)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(500, 240, 101, 20))
        font = QtGui.QFont()
        font.setPointSize(7)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.toplam = 0
        self.fisid = 1
        self.tarih = datetime.now()
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Vekon Excel Harcama Oluşturucu"))
        self.pushButton.setText(_translate("MainWindow", "EKLE"))
        self.pushButton_2.setText(_translate("MainWindow", "ÇIKIŞ"))
        self.label.setText(_translate("MainWindow", "İŞ NUMARASI "))
        self.pushButton_3.setText(_translate("MainWindow", "KAYDET"))
        self.label_3.setText(_translate("MainWindow", "    1          2        3     4     5        6          7       8        9"))
        self.label_4.setText(_translate("MainWindow", "MERT FINDIKLI"))
        self.pushButton.setShortcut(_translate("MainWindow", "Return"))
    def exit(self):
        exit()
    def ileri(self):
        if self.toplam == 0:
                cunsor.execute("Insert into ana (isadi) Values('{}')".format(self.lineEdit.text().upper()))
                self.label.setText("FİRMA ADI ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......",1000)
        elif self.toplam == 1:
                cunsor.execute("update ana set firma = '{}'".format(self.lineEdit.text().upper()))
                self.label.setText("AVANSI ALAN ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
        elif self.toplam == 2:
                cunsor.execute("update  ana set avansialan = '{}'".format(self.lineEdit.text().upper()))
                self.label.setText("VERİLEN AVANS TUTARI ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
        elif self.toplam == 3:
                cunsor.execute("update  ana set verilen = '{}'".format(self.lineEdit.text().upper()))
                self.label.setText("FİŞ TARİHİ ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
        elif self.toplam == 4:
                cunsor.execute("Insert into FİRMA (id,tarih) Values('{}','{}')".format(self.fisid,
                                                                               self.lineEdit.text().upper()))
                self.label.setText("BELGE NO ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
        elif self.toplam == 5:
                cunsor.execute("update  FİRMA set belge = '{}' where id = '{}'".format(self.lineEdit.text().upper(),
                                                                                        self.fisid))
                self.label.setText("HARCAMA FİRMA ADI ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
        elif self.toplam == 6:
                cunsor.execute("update FİRMA set FİRMA = '{}' where id = '{}'".format(self.lineEdit.text().upper(),
                                                                                       self.fisid))
                self.label.setText("KİŞİ SAYISI ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
        elif self.toplam == 7:
                cunsor.execute("update FİRMA set Kişi = '{}' where id = '{}'".format(self.lineEdit.text().upper(),
                                                                                      self.fisid))
                self.label.setText("ÜSTEKİLERDEN HANGİSİ  ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
        elif self.toplam == 8:
                cunsor.execute("update FİRMA set gider = '{}' where id = '{}'".format(self.lineEdit.text().upper(),
                                                                                       self.fisid))
                self.label.setText("FİYAT NEDİR ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
        elif self.toplam == 9:
                cunsor.execute("update FİRMA set toplam = '{}' where id = '{}'".format(self.lineEdit.text().upper(),
                                                                                        self.fisid))
                self.label.setText("FİŞ TARİHİ  ")
                self.lineEdit.clear()
                self.toplam -= 5
                self.fisid += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
        baglanti.commit()

    def kaydet(self):
        import shutil
        shutil.copy("gerekli/test.xlsx", "./")
        self.dosya_ac = openpyxl.load_workbook("./test.xlsx")
        self.sayfa_ac = self.dosya_ac["Vekon Harcama"]
        self.sayfa_sayisi = self.sayfa_ac.max_row
        cunsor.execute("SELECT * FROM ana")
        self.islem = cunsor.fetchall()
        cunsor.execute("SELECT * FROM FİRMA")
        self.islem1 = cunsor.fetchall()
        self.toplam = 10
        self.toplamfiyat = 0
        for i in self.islem:
            self.sayfa_ac.cell(1, 3, value="{}".format(i[0]))
            self.sayfa_ac.cell(2, 3, value="{}".format(i[1]))
            self.sayfa_ac.cell(3, 3, value="{}".format(i[2]))
            self.sayfa_ac.cell(5, 3, value="{}".format(i[3]))
            self.sayfa_ac.cell(45, 2, value="{}".format(i[2]))
        for e in self.islem1:
            self.sayfa_ac.cell(self.toplam, 1, value="{}".format(e[1]))
            self.sayfa_ac.cell(self.toplam, 2, value="{}".format(e[2]))
            self.sayfa_ac.cell(self.toplam, 3, value="{}".format(e[3]))
            self.sayfa_ac.cell(self.toplam, 4, value="{}".format(e[4]))
            if e[5] == "1":
                self.sayfa_ac.cell(self.toplam, 5, value="{}".format("X"))
            elif e[5] == "2":
                self.sayfa_ac.cell(self.toplam, 6, value="{}".format("X"))
            elif e[5] == "3":
                self.sayfa_ac.cell(self.toplam, 7, value="{}".format("X"))
            elif e[5] == "4":
                self.sayfa_ac.cell(self.toplam, 8, value="{}".format("X"))
            elif e[5] == "5":
                self.sayfa_ac.cell(self.toplam, 9, value="{}".format("X"))
            elif e[5] == "6":
                self.sayfa_ac.cell(self.toplam, 10, value="{}".format("X"))
            elif e[5] == "7":
                self.sayfa_ac.cell(self.toplam, 11, value="{}".format("X"))
            elif e[5] == "8":
                self.sayfa_ac.cell(self.toplam, 12, value="{}".format("X"))
            elif e[5] == "9":
                self.sayfa_ac.cell(self.toplam, 13, value="{}".format("X"))
            self.sayfa_ac.cell(self.toplam, 16, value="{}".format(float(e[6])))
            self.toplam += 1
            self.toplamfiyat += e[6]
        self.sayfa_ac.cell(45, 1, value="{}".format(self.tarih.date()))
        self.sayfa_ac.cell(41, 16, value="{}".format(float(self.toplamfiyat)))
        self.dosya_ac.save("./test.xlsx")
        os.rename("test.xlsx", "{}_{}_{}_HARCAMA_BELGESİ_{}.xlsx".format((str(i[2]).strip("\n\n")),(str(i[0]).strip("\n\n")),(str(i[1]).strip("\n\n")),self.tarih.date()))
        cunsor.execute("delete from FİRMA")
        cunsor.execute("delete from ana")
        baglanti.commit()
import resim_rc

if __name__ == "__main__":
        import sys
        app = QtWidgets.QApplication(sys.argv)
        pencere = QtWidgets.QMainWindow()
        ui = Ui_MainWindow()
        ui.setupUi(pencere)
        pencere.show()
        sys.exit(app.exec_())
